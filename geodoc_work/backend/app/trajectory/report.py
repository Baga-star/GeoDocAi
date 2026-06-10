from __future__ import annotations

from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, ScatterChart, Series
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.trajectory.models import DeviationRow, SeparationRow, SourceProvenance, TrajectoryPoint

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_sheet(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for column in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(max(max_len + 2, 12), 42)


def _append_dicts(ws, headers: list[str], rows: Iterable[dict]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    _style_sheet(ws)


def _source_rows(sources: list[SourceProvenance]) -> list[dict]:
    return [
        {
            "document_id": s.document_id,
            "document_name": s.document_name,
            "page": s.page,
            "artifact_id": s.artifact_id,
            "table_title": s.table_title,
            "row_index": s.row_index,
            "raw": str(s.raw)[:300],
        }
        for s in sources
    ]


def build_excel_report(
    *,
    project_name: str,
    well_name: str,
    survey_rows: list[dict],
    actual_points: list[TrajectoryPoint],
    design_points: list[TrajectoryPoint],
    deviation_rows: list[DeviationRow],
    separation_rows: list[SeparationRow],
    sources: list[SourceProvenance],
    warnings: list[str],
) -> bytes:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Field", "Value"])
    summary.append(["Project", project_name])
    summary.append(["Well", well_name])
    summary.append(["Depth sign", "TVD positive downward"])
    summary.append(["Survey rows", len(survey_rows)])
    summary.append(["Computed actual points", len(actual_points)])
    summary.append(["Computed design points", len(design_points)])
    summary.append(["Deviation rows", len(deviation_rows)])
    summary.append(["Separation rows", len(separation_rows)])
    summary.append(["Image charts", "Not embedded in Phase 1; JSON/worksheet data is authoritative."])
    _style_sheet(summary)

    ws = wb.create_sheet("SurveyData")
    _append_dicts(ws, ["md", "inc", "azi", "magnetic_declination", "approved", "source"], survey_rows)

    ws = wb.create_sheet("ComputedCoordinates")
    _append_dicts(
        ws,
        ["layer", "md", "inc", "azi", "tvd", "northing", "easting", "vertical_section"],
        [p.model_dump() for p in actual_points],
    )

    plan = wb.create_sheet("Plan")
    _append_dicts(plan, ["md", "easting", "northing", "layer"], [{"md": p.md, "easting": p.easting, "northing": p.northing, "layer": p.layer} for p in actual_points])
    if len(actual_points) >= 2:
        chart = ScatterChart()
        chart.title = "Plan: easting/northing"
        chart.x_axis.title = "Easting, m"
        chart.y_axis.title = "Northing, m"
        xvalues = Reference(plan, min_col=2, min_row=2, max_row=len(actual_points) + 1)
        yvalues = Reference(plan, min_col=3, min_row=2, max_row=len(actual_points) + 1)
        chart.series.append(Series(yvalues, xvalues, title="actual"))
        plan.add_chart(chart, "F2")

    profile = wb.create_sheet("Profile")
    _append_dicts(profile, ["md", "vertical_section", "tvd", "layer"], [{"md": p.md, "vertical_section": p.vertical_section, "tvd": p.tvd, "layer": p.layer} for p in actual_points])
    if len(actual_points) >= 2:
        chart = ScatterChart()
        chart.title = "Profile: vertical section/TVD"
        chart.x_axis.title = "Vertical section, m"
        chart.y_axis.title = "TVD, m"
        xvalues = Reference(profile, min_col=2, min_row=2, max_row=len(actual_points) + 1)
        yvalues = Reference(profile, min_col=3, min_row=2, max_row=len(actual_points) + 1)
        chart.series.append(Series(yvalues, xvalues, title="actual"))
        profile.add_chart(chart, "F2")

    design = wb.create_sheet("Design")
    _append_dicts(design, ["layer", "md", "inc", "azi", "tvd", "northing", "easting"], [p.model_dump() for p in design_points])

    deviation = wb.create_sheet("Deviation")
    _append_dicts(deviation, ["md", "tvd", "northing", "easting", "nearest_design_md", "distance_m", "delta_tvd", "delta_northing", "delta_easting"], [r.model_dump() for r in deviation_rows])

    separation = wb.create_sheet("Separation")
    _append_dicts(separation, ["well_a_id", "well_a_name", "well_b_id", "well_b_name", "min_distance_m", "md_a", "md_b", "method"], [r.model_dump() for r in separation_rows])

    src = wb.create_sheet("Sources")
    _append_dicts(src, ["document_id", "document_name", "page", "artifact_id", "table_title", "row_index", "raw"], _source_rows(sources))

    warn = wb.create_sheet("Warnings")
    warn.append(["warning"])
    for item in warnings:
        warn.append([item])
    _style_sheet(warn)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
